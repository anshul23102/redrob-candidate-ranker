#!/usr/bin/env python3
"""Convert the submission CSV to XLSX (identical rows/columns).

The bundle's validator requires CSV, but the portal upload widget accepts
excel/spreadsheet - we ship both, byte-equivalent in content.

Usage: python scripts/csv_to_xlsx.py the_last_commit.csv
"""
import csv, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

src = sys.argv[1] if len(sys.argv) > 1 else "the_last_commit.csv"
dst = src.rsplit(".", 1)[0] + ".xlsx"

wb = Workbook()
ws = wb.active
ws.title = "ranking"
with open(src, encoding="utf-8", newline="") as f:
    for row in csv.reader(f):
        ws.append(row)
# rank/score as numbers (not text), uniform 6-decimal display for score
for r in range(2, ws.max_row + 1):
    ws.cell(r, 2).value = int(ws.cell(r, 2).value)
    sc = ws.cell(r, 3)
    sc.value = float(sc.value)
    sc.number_format = "0.000000"
    ws.cell(r, 4).alignment = Alignment(vertical="top")
# presentation: bold frozen header, sane column widths
for c in ws[1]:
    c.font = Font(bold=True)
ws.freeze_panes = "A2"
for col, w in zip("ABCD", (16, 7, 11, 110)):
    ws.column_dimensions[col].width = w
wb.save(dst)
print(f"wrote {dst} ({ws.max_row - 1} data rows)")
